## File naming and formatting
# testing
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
#goes to other file
from openpyxl import Workbook
import pandas as pd
from openpyxl.drawing.image import Image
from pathlib import Path
import statistics
import json
import shutil

import os
from dotenv import load_dotenv

dotenv_path = "/Users/blayton/Documents/env/.env"
load_dotenv(dotenv_path)

def excel_file_path (customer_name, purpose_name):
    JSON_FILE_PATH = os.getenv("JSON_FILE_PATH", "default")
    REPORT_OUTPUT_PATH = os.getenv("REPORT_OUTPUT_PATH", "default")

    if JSON_FILE_PATH != "default":
        with open(JSON_FILE_PATH,"r") as file:
            config = json.load(file)

    customer_name = config["buca_payers"]
    purpose_name = config["purpose_name"]
    title_page_data_used = config["title_page_data_used"]

    """Intakes the customer name and purpose with the current time stamp, sets file path. If customer parent folder does not exist, it makes one."""
    current_date = datetime.now().strftime('%Y_%m_%d')
    file_name = f"{customer_name.lower()}_{purpose_name.lower()}_{current_date}.xlsx"
    file_path = Path(f"{REPORT_OUTPUT_PATH}{customer_name.title()}/{file_name}")
    file_path.parent.mkdir(parents=True,exist_ok=True)
    return file_path

#def adjust_column_widths(worksheet, dataframe):
   # """Sets column widths"""
   # for col_idx, column in enumerate(dataframe.columns, start=1):
    #    max_length = max((len(str(value)) for value in dataframe[column].values), default=10)
    #    max_length = max(max_length, len(str(column)))  # Ensure column name fits
   #     worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 10 

def adjust_column_widths_max(ws):
    """Adjust column widths based on content in the worksheet."""
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # Adjust width for the column name/header
        col_letter = get_column_letter(col_idx)
        column_header = ws[f"{col_letter}1"].value
        if column_header:
            max_length = max(max_length, len(str(column_header)))
        # Apply the width with some padding
        ws.column_dimensions[col_letter].width = max_length + 2

def adjust_column_widths_fixed(ws):
    """Set a fixed column width of 25 with padding for all columns."""
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        # Set a fixed column width of 25 plus some padding (e.g., 2)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25 + 2 

def adjust_column_widths_median(ws):
    """
    Adjust column widths based on the median width of the content in each column.
    """
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        # Collect lengths of all non-empty cell values in the column
        lengths = [
            len(str(cell.value)) for cell in col_cells if cell.value is not None
        ]
        # Add the column header length, if it exists
        col_letter = get_column_letter(col_idx)
        column_header = ws[f"{col_letter}1"].value
        if column_header:
            lengths.append(len(str(column_header)))

        # Set the column width based on the median length
        if lengths:  # Avoid empty columns
            median_length = statistics.median(lengths)
            ws.column_dimensions[col_letter].width = median_length + 2  # Add padding


def apply_style(cell, font, fill, border, alignment):
    """Sets style for Excel sheet"""
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment

def title_page_formatting():
    title_style = {
                'font': Font(name='Arial', color="02363D", bold=True, size=14),
                'fill': PatternFill(start_color="F7FBFB",end_color="F7FBFB",fill_type="solid"),
                'border': Border(),
                'alignment': Alignment(horizontal="left")
            }

    sub_title_style = {
        'font': Font(name='Arial', color="02363D", bold=True, size=12),
        'fill': PatternFill(start_color="F7FBFB",end_color="F7FBFB"),
        'border': Border(),
        'alignment': Alignment(horizontal="left")
    }

    header_border_side = Side(border_style="thin", color="02363D") 

    header_style = {
        'font': Font(name='Arial', color="FFFFFF", size = 12),
        'fill': PatternFill(start_color="02363D", end_color="02363D", fill_type="solid"),
        'border': Border(left=header_border_side, right=header_border_side,
                        top=header_border_side, bottom=header_border_side),
        'alignment': Alignment(horizontal="left")
    }

    center_style = {
        'font': Font(name='Arial', color="02363D", size = 12),
        'fill': PatternFill(start_color="FFFFFF", end_color="FFFFFF"),
        'border': Border(left=header_border_side, right=header_border_side,
                        top=header_border_side, bottom=header_border_side),
        'alignment': Alignment(horizontal="left")
    }


    return title_style, header_style, center_style


def write_title_boiler(ws, customer_name):
    title_style, header_style, center_style= title_page_formatting()
    #Set background color
    for row in ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=200):
        for cell in row:
            apply_style(cell, title_style['font'], title_style['fill'], title_style['border'], title_style['alignment'])

    #Add TQ logo
    img = Image("tq_logo.png")
    img.width = 500
    img.height = 80
    ws.add_image(img, 'A1')

    # Write Title data
    cell = ws.cell(row=8, column=1, value="TQ Negotiated Rate Analysis")
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
    apply_style(cell, title_style['font'], title_style['fill'], title_style['border'], title_style['alignment'])

    # Subtitle data
    cell = ws.cell(row=9, column=1, value=customer_name)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
    apply_style(cell, title_style['font'], title_style['fill'], title_style['border'], title_style['alignment'])

    # Write "extract is based off" and table of contents header
    cell = ws.cell(row=11, column=1, value=f"Extract is based on {title_page_data_used} data.")
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=2)
    cell = ws.cell(row=13, column=1, value="Table of Contents")
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=2)


def write_title_page_tables(ws, x_df,last_row):
    headers = list(x_df.columns)
    title_style, header_style, center_style = title_page_formatting()

    if last_row == 0:
        header_row = 14

        for col_idx, column in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=column)
            apply_style(cell, header_style['font'], header_style['fill'], header_style['border'], header_style['alignment'])

        for row_idx, row in enumerate(x_df.values, start=header_row+1):
            for col_idx, column in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=column)
                apply_style(cell, center_style['font'], center_style['fill'], center_style['border'], center_style['alignment'])
        
        last_row = row_idx

        return last_row

    else: 
        header_row = last_row + 2

        for col_idx, column in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=column)
            apply_style(cell, header_style['font'], header_style['fill'], header_style['border'], header_style['alignment'])

        # table rows
        for row_idx, row in enumerate(x_df.values, start=header_row+1):
            for col_idx, column in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=column)
                apply_style(cell, center_style['font'], center_style['fill'], center_style['border'], center_style['alignment'])

        last_row = row_idx

        return last_row
    # Adjust column widths based on the DataFrame
        #adjust_column_widths(ws, x_df)

    
    

def data_page_formatting():
    # Create different styles
    border_side = Side(style='thin', color='FFFFFF')
    header_style = {
        'font': Font(bold=True, color="FFFFFF", name='Arial', size=12),
        'fill': PatternFill(start_color="02363D", end_color="02363D", fill_type="solid"),
        'border': Border(left=border_side, right=border_side, top=border_side, bottom=border_side),
        'alignment': Alignment(horizontal='center', vertical='center')
    }

    data_style = {
        'font': Font(name='Arial', size=12),
        'fill': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'border': Border(left=border_side, right=border_side, top=border_side, bottom=border_side),
        'alignment': Alignment(horizontal='left', vertical='center')
    }

    return header_style, data_style


def data_page_write(ws, x_df):
    header_style, data_style = data_page_formatting()

    # Flatten MultiIndex columns if needed
    if isinstance(x_df.columns, pd.MultiIndex):
        x_df.columns = [" | ".join(map(str, col)) for col in x_df.columns]

    if isinstance(x_df.index, pd.MultiIndex):
        x_df = x_df.reset_index()

    # Flatten lists in DataFrame
    for col in x_df.columns:
        if x_df[col].apply(lambda x: isinstance(x, list)).any():
            x_df[col] = x_df[col].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)

    # Write headers
    for col_idx, column in enumerate(x_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        apply_style(cell, header_style['font'], header_style['fill'], header_style['border'], header_style['alignment'])

    # Write data
    for row_idx, row in enumerate(x_df.values, start=2):
        for col_idx, column in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=column)
            apply_style(cell, data_style['font'], data_style['fill'], data_style['border'], data_style['alignment'])
            if row_idx % 2 == 0:  # Color even rows
                cell.fill = PatternFill(start_color="E7F3F2", end_color="E7F3F2", fill_type="solid")
    
    #adjust_column_widths(ws, x_df)



def special_write_function(title_data_frames, title_sheet_name, data_df_dict, customer_name):
    """Takes the title data frames: list , title sheet name and writes the title page. Takes the data_df_dict to write each tab of the data"""
    
    # Initialize workbook
    wb = Workbook()

    compl_row = 0
    for summary_df in title_data_frames:

        if compl_row == 0:
            # Initialize worksheet
            ws = wb.create_sheet(title_sheet_name)
            write_title_boiler(ws,customer_name)
            compl_row = write_title_page_tables(ws,summary_df,last_row=0)
    
        else:
            ws = wb[ title_sheet_name]
            compl_row = write_title_page_tables(ws, summary_df, last_row=compl_row)

        
    for sheet_name, df in data_df_dict.items():
        ws = wb.create_sheet(sheet_name)
        data_page_write(ws,df)

    if "Sheet" in wb.sheetnames:
        del wb['Sheet']

    if title_sheet_name in wb.sheetnames:
        ws = wb[title_sheet_name]
        adjust_column_widths_max(ws) 
    
    rates_sheet_name = next(reversed(data_df_dict.keys()))  # Ensure it's getting the last key explicitly

    if rates_sheet_name in wb.sheetnames:
        ws = wb[rates_sheet_name]  # Make sure the correct worksheet is accessed
        adjust_column_widths_median(ws)

    # Process other sheets
    keys = list(data_df_dict.keys())  # Convert keys to a list for indexed access
    for index, key in enumerate(keys[:-1]):  # Ignore the last key
        sheet_name = key  # Correctly set the sheet name from the key
        if sheet_name in wb.sheetnames:  # Validate the sheet name exists
            ws = wb[sheet_name]
            adjust_column_widths_max(ws)  # Adjust column widths for other sheets
            print(f"Processing key: {key}, value: {sheet_name}")

def empty_directory(directory_path):
    # Check if the directory exists
    if os.path.exists(directory_path):
        # Iterate through all the files and subdirectories in the directory
        for item in os.listdir(directory_path):
            item_path = os.path.join(directory_path, item)

            if os.path.isdir(item_path):
                # If it's a directory, remove it recursively
                shutil.rmtree(item_path)
                print(f"Deleted directory: {item_path}")
            else:
                # If it's a file, remove it
                os.remove(item_path)
                print(f"Deleted file: {item_path}")
    else:
        print(f"Directory {directory_path} does not exist")

   
def cool_write_function(customer_abbreviation, output_dir, sql_path, df):
    """Takes the customer abbreviation and sql_path to use in the file name, uses the df to populate the page"""

    # Initialize workbook
    wb = Workbook()
    ws = wb.active 

    data_page_write(ws, df)
    adjust_column_widths_fixed(ws)

    base_file_name = os.path.basename(sql_path).replace('.sql', '')
    current_date = datetime.now()
    formatted_date = current_date.strftime('%Y%m%d')

    os.makedirs(output_dir, exist_ok=True)

    file_path_string = f"cc_{customer_abbreviation}_{base_file_name}_{formatted_date}".upper()
    file_path = os.path.join(output_dir, f"{file_path_string}.xlsx")
    wb.save(filename=file_path)

    print(f"File saved to {file_path}")

def export_to_csv(customer_abbreviation, sql_path, df, logger, quote_setting, escapechar):
    """Takes a df and exports as csv, converts all columns with external id in the name to strings to maintain """

    try: 
        base_file_name = os.path.basename(sql_path).replace('.sql', '')
        current_date = datetime.now()
        formatted_date = current_date.strftime('%Y%m%d')

        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)

        file_path_string = f"cc_{customer_abbreviation}_{base_file_name}_{formatted_date}".upper()
        file_path = os.path.join(output_dir, f"{file_path_string}.csv")

        df.to_csv(file_path, index=False, quoting=quote_setting, escapechar=escapechar)
        logger.info(f"{file_path} saved sucessfully in {output_dir}!")

    except Exception as e:
        logger.error(f"Error exporting CSV: {str(e)}")